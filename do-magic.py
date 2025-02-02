import os
import warnings
from tqdm import tqdm
import string
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font

warnings.simplefilter("ignore")

def set_border(ws, cell_range):
    thin = openpyxl.styles.Side(border_style="thin", color="757171")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.comment = None
            # cell.alignment = Alignment(wrap_text=True)


def set_header(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = openpyxl.styles.PatternFill(start_color="ffffff", fill_type="solid")
            cell.font = cell.font.copy(color="000000")
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')


def check_end(ws, start, col):
    no_end = True
    start = 8
    while no_end:
        if ws['{}{}'.format(col, start)].value:
            pass
        else:
            no_end = False
            return start - 1
        start += 1


def check_max_col(ws):
    no_end = True
    i = 8
    while no_end:
        if not ws['{}4'.format(get_column_letter(i))].value:
            no_end = False
        i += 1
    return get_column_letter(i - 2)


def check_start_a(ws):
    for i in range(5, 19):
        if ws['a{}'.format(i)].value:
            return (i)


def check_start_f(ws):
    col_range = list(string.ascii_lowercase)
    for i, n in enumerate(col_range):
        if ws['{}2'.format(n)].value:
            return col_range[i], col_range[i - 1], col_range[i + 1]

def check_for_hide_colums(n, e, ws):
    to_hide = []
    col_range = list(string.ascii_lowercase)
    for i, col in enumerate(col_range):
        if '[Attr' in str(ws['{}{}'.format(col,n)].value):
            to_hide.append(col)
    return(to_hide)


def hide_cols(to_hide, ws):
    for col in to_hide:
        ws.column_dimensions[col.upper()].hidden = True       
    return ws


def check_for_hide_rows(n, m, ws):
    to_hide = []
    for row in range(1, n):
        if '[Attr' in str(ws['{}{}'.format(m, row)].value):
            to_hide.append(row)
    return to_hide


def hide_rows(to_hide, ws):
    for row in to_hide: 
        ws.row_dimensions[row].hidden = True     
    return ws    


def format_first_type(ws):
    default_column_width = 25
    for i in list(string.ascii_lowercase):
        ws.column_dimensions[i].width = default_column_width
    n = check_start_a(ws) # start of first block
    m, e , x = check_start_f(ws) # m = start column of second block; e = end column of first block; x = freeze column
    end = check_end(ws, 8, 'A') # end of first block
    col = check_max_col(ws) # end column
    to_hide = check_for_hide_colums(n, e, ws)
    ws = hide_cols(to_hide, ws)
    to_hide = check_for_hide_rows(n, m, ws)
    ws = hide_rows(to_hide, ws)    
    set_border(ws, 'A{}:{}{}'.format(n, e, end))
    set_border(ws, '{}1:{}{}'.format(m, col, end))
    set_header(ws, 'A{}:{}{}'.format(n, e, n))
    set_header(ws, '{}1:{}{}'.format(m, m, n - 1))
    ws['A1'].alignment = Alignment(horizontal='left', vertical = 'top', wrap_text=True)
    ws['A1'].font = Font(size="9", bold=True, italic=True)
    ws.merge_cells('A1:{}{}'.format(e, n -1))
    ws.freeze_panes = '{}{}'.format(x, n+1)
    return ws


def next_alpha(s):
    return chr((ord(s.upper())+1 - 65) % 26 + 65)


def format_information_result(ws, last_tab):
    default_column_width = 20
    end = check_end(ws, 1, 'B') + 1
    for i in range(1, end):
        if ws['a1'].value == "ID":
            ws['a{}'.format(i)].alignment = Alignment(horizontal='center')
        ws['{}{}'.format(last_tab, i)].alignment = Alignment(horizontal='center')

    for i in list(string.ascii_lowercase):
        ws.column_dimensions[i].width = default_column_width
        if ws['a1'].value == "ID" and i == "a":
            ws.column_dimensions[i].width = "5"
        if ws['{}1'.format(i)].value == "Status" :
            ws.column_dimensions[i].width = "40"
        if ws['{}1'.format(i)].value == "Condition" :
            ws.column_dimensions[i].width = "28"
        elif ws['{}1'.format(i)].value == "Process" or ws['{}1'.format(i)].value == "C/E":
            ws.column_dimensions[i].width = "10"
            for row in range(1, end):
                ws['{}{}'.format(i, row)].alignment = Alignment(horizontal='center')

    ws.column_dimensions[last_tab].width = 5
    ws.merge_cells('{}1:{}1'.format(last_tab, next_alpha(last_tab)))
    set_header(ws, 'A1:{}1'.format(last_tab))
    ws.freeze_panes = ws['a2']
    return ws


def format_status_table(ws, last_tab):
    default_column_width = 37
    ws.merge_cells('b1:c1')
    ws.merge_cells('e1:f1')
    if str(ws['b1'].value) == 'None':
        ws['b1'].value = ws['a1'].value
        ws['b1'].font = Font(bold=True, name='Dialog.bold')
        ws['a1'].value = ""
    set_header(ws, 'A1:{}2'.format(last_tab))
    ws.freeze_panes = ws['a3']
    for i in list(string.ascii_lowercase):
        ws.column_dimensions[i].width = default_column_width
        if i == "a" : ws.column_dimensions[i].width = "12"

        if ("IDENTIFIER" in str(ws['{}2'.format(i)].value) or ws['{}1'.format(i)].value == "to") and (str(ws['{}3'.format(i)].value).isdigit() or str(ws['{}3'.format(i)].value) == 'None') :
            ws.column_dimensions[i].width = "4"
            ws['{}2'.format(i)].alignment = Alignment(horizontal='left')
            ws['{}1'.format(i)].alignment = Alignment(horizontal='left')
    return ws


def find_last_tab_2(ws):
    col_range = list(string.ascii_lowercase)
    for c in col_range:
        if "Document Status" in str(ws['{}1'.format(c)].value) and c != "a" :
            return next_alpha(c)
    return "f"


def find_last_tab(ws):
    col_range = list(string.ascii_lowercase)
    for c in col_range:
        if  "CERTEX" in str(ws['{}1'.format(c)].value):
            return c
    return "d"

def column_letters():
    col_range = list(string.ascii_lowercase)
    new_list = list(string.ascii_lowercase)

    for c in col_range:
        for letter in col_range:
            new_list.append(c + letter)
    return new_list


def format_information_result_reacap(ws):
    ws.freeze_panes = ws['h3']
    ws.sheet_view.zoomScale = 70
    ws['a1'].alignment = Alignment(horizontal='center', wrap_text = True)
    ws['c2'].alignment = Alignment(horizontal='center', wrap_text = True)
    ws['d2'].alignment = Alignment(horizontal='center', wrap_text = True)
    ws['e2'].alignment = Alignment(horizontal='center', wrap_text = True)
    ws['f2'].alignment = Alignment(horizontal='center', wrap_text = True) 
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 51
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['H'].width = 5
    ws.column_dimensions['K'].width = 5
    ws.column_dimensions['R'].width = 5
    ws.column_dimensions['U'].width = 5
    ws.column_dimensions['X'].width = 5
    ws.column_dimensions['AA'].width = 5
    ws.column_dimensions['AD'].width = 5
    ws.column_dimensions['AF'].width = 5
    ws.column_dimensions['B'].width = 43
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 17
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 26
    ws.column_dimensions['L'].width = 40
    ws.column_dimensions['M'].width = 26
    ws.column_dimensions['N'].width = 29
    ws.column_dimensions['O'].width = 26
    ws.column_dimensions['P'].width = 12
    ws.column_dimensions['Q'].width = 26
    ws.column_dimensions['S'].width = 40
    ws.column_dimensions['T'].width = 26
    ws.column_dimensions['V'].width = 40
    ws.column_dimensions['W'].width = 26
    ws.column_dimensions['Y'].width = 40
    ws.column_dimensions['Z'].width = 26
    ws.column_dimensions['AB'].width = 40
    ws.column_dimensions['AC'].width = 26
    ws.column_dimensions['AD'].width = 40
    ws.column_dimensions['AE'].width = 26
    ws.column_dimensions['AG'].width = 29
    ws.column_dimensions['AH'].width = 26
    ws.merge_cells('A1:B1')
    ws.merge_cells('C1:F1')
    ws.merge_cells('H1:J1')
    ws.merge_cells('K1:M1')
    ws.merge_cells('N1:O1')
    ws.merge_cells('P1:Q1')
    ws.merge_cells('R1:T1')
    ws.merge_cells('U1:W1')
    ws.merge_cells('K1:M1')
    ws.merge_cells('X1:Z1')
    ws.merge_cells('K1:M1')
    ws.merge_cells('AA1:AC1')
    ws.merge_cells('K1:M1')
    ws.merge_cells('AD1:AE1')
    ws.merge_cells('AF1:AH1')

    return ws

def main():
    done_folder = "done"
    work_folder = "work"
    full_work_folder = os.path.join(os.path.dirname(os.path.realpath(__file__)), work_folder, "")
    dir_list = []

    for file in  os.listdir(full_work_folder):
        if file.endswith(".xlsx"):
            dir_list.append(file)

    print("Formatting all {} .xlsx files found in {}".format(len(dir_list), full_work_folder))   

    for filename in tqdm(dir_list, desc ="Work done: "):
        if "Transformation Table" in filename and "Status" not in filename:
            wb = openpyxl.load_workbook(os.path.join(full_work_folder, filename))
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                ws.sheet_view.zoomScale = 70
                ws = format_first_type(ws)
            wb.save(os.path.join(os.path.dirname(os.path.realpath(__file__)), done_folder, filename))
        elif "Information Result" in filename:
            wb = openpyxl.load_workbook(os.path.join(full_work_folder, filename))
            for ws_name in wb.sheetnames:
                if "Recap" not in ws_name:
                    ws = wb[ws_name]
                    ws.sheet_view.zoomScale = 70
                    ws = format_information_result(ws, find_last_tab(ws))                   
                else:
                    ws = wb[ws_name]
                    ws = format_information_result_reacap(ws)
            wb.save(os.path.join(os.path.dirname(os.path.realpath(__file__)), done_folder, filename))
        elif "Status Transformation Table" in filename:
            wb = openpyxl.load_workbook(os.path.join(full_work_folder, filename))
            for ws_name in wb.sheetnames:
                # if "Statuses" not in ws_name:
                ws = wb[ws_name]
                ws.sheet_view.zoomScale = 70
                ws = format_status_table(ws, find_last_tab_2(ws))
            wb.save(os.path.join(os.path.dirname(os.path.realpath(__file__)), done_folder, filename))

        elif "Spreadsheet Rules Table" in filename:
            wb = openpyxl.load_workbook(os.path.join(full_work_folder, filename))
            col_range = column_letters()
            wb[wb.sheetnames[0]].column_dimensions['B'].width = 63.44
            wb[wb.sheetnames[0]].freeze_panes = wb[wb.sheetnames[0]]['d6']
            wb[wb.sheetnames[1]].freeze_panes = wb[wb.sheetnames[1]]['c6']
            wb[wb.sheetnames[2]].freeze_panes = wb[wb.sheetnames[2]]['c6']
            wb[wb.sheetnames[3]].freeze_panes = wb[wb.sheetnames[3]]['c5']
            wb[wb.sheetnames[4]].freeze_panes = wb[wb.sheetnames[4]]['f6']
            wb[wb.sheetnames[5]].freeze_panes = wb[wb.sheetnames[5]]['c6']
            wb[wb.sheetnames[6]].freeze_panes = wb[wb.sheetnames[6]]['c6']
            wb[wb.sheetnames[7]].freeze_panes = wb[wb.sheetnames[7]]['d6']

            for name in wb.sheetnames:
                ws = wb[name]
                ws.sheet_view.zoomScale = 70
                ws.row_dimensions[1].height = 30 
                ws.row_dimensions[2].height = 157.1
                ws.row_dimensions[4].height = 157.1 
                         

                for column in col_range:
                    ws['{}1'.format(column)].font = Font(size = '14', bold = True, name='Dialog.bold')
                    ws['{}2'.format(column)].font = Font(size = '10', name='Dialog.plain')
                    ws['{}4'.format(column)].font = Font(size = '10', name='Dialog.plain')
                    ws['{}2'.format(column)].alignment = Alignment(textRotation = 90, horizontal='left', wrap_text = True)
                    ws['{}4'.format(column)].alignment = Alignment(textRotation = 90, horizontal='left', wrap_text = True)

            wb.save(os.path.join(os.path.dirname(os.path.realpath(__file__)), done_folder, filename))

 
if __name__ == '__main__':
    main()